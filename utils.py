from openpyxl import load_workbook

wb = load_workbook("Materiamidia.xlsx")

dicionario = [
        "Banco do Brasil", "Banco do Nordeste", "Banco da Amazônia", "Caixa Econômica Federal", "Banco Central", "Casa da Moeda", "Comissão de Valores Mobiliários", "Susep", "BNDES", "BB", "BNB", "BASA", "CEF", "CMB", "ABFG", "AGROS","CAPEF","CAPESESP","CENTRUS","CERES","CIBRIUS","CIFRAO","ELETROS","ELOS","FACHESF","FAPES","FIOPREV","FIPECQ","FUNCEF","FUNPRESP-EXE","FUNPRESP-JUD","FUSESC","GEIPREV","INFRAPREV","NUCLEOS","PETROS","POSTALIS","PREVBEP","PREVDATA","PREVI","PREVINORTE","REAL GRANDEZA","REFER","SAO FRANCISCO","SERPROS","SIAS","CAPAF","PORTUS","TCU"
  ]




def organizacao_midias(item):
    if item == "AGROS" or item == "CAPEF" or item == "CAPESESP" or item == "CENTRUS" or item == "CERES" or  item == "CIBRIUS" or item == "CIFRAO" or item == "ELETROS" or item == "ELOS" or item == "FACHESF" or item == "FAPES" or item == "FIOPREV" or item == "FIPECQ" or item == "FUNCEF" or item == "FUNPRESP-EXE" or item == "FUNPRESP-JUD" or item == "FUSESC" or item == "GEIPREV" or item == "INFRAPREV" or item == "NUCLEOS" or item == "PETROS" or item == "POSTALIS" or item == "PREVBEP" or item == "PREVDATA" or item == "PREVI" or item == "PREVINORTE" or item == "REAL GRANDEZA" or item == "REFER" or item == "SAO FRANCISCO" or item == "SERPROS" or item == "SIAS" or item == "CAPAF" or item == "PORTUS":
      worksheet="DT1"
    elif item == "BB":
      worksheet="Banco do Brasil"
    elif item == "Comissão de Valores Mobiliários":
      worksheet="CVM"
    elif item == "BNB":
      worksheet="Banco do Nordeste"
    elif item == "BASA":
      worksheet="Banco da Amazônia"
    elif item == "CEF":
      worksheet="Caixa Econômica Federal"
    elif item == "Superintendência de Seguros Privados":
      worksheet="Susep"
    elif item == "Associação Brasileira de Fundos Garantidores":
      worksheet="Supep" 
    else:
       worksheet=item
    return worksheet

def compilacao(wb):
  wb = load_workbook("Materiamidia.xlsx")
  try:
    compilado = wb["COMPILADO"]
  except:
    wb.create_sheet("COMPILADO")
    compilado = wb["COMPILADO"]
    titulo = ("Titulo","Link","Midia","Data", "Clientela")
    compilado.append(titulo)
  for item in wb.sheetnames:
    ws = wb[item]
    for row in ws.iter_rows(min_row=2,max_col=5, max_row=ws.max_row):
      try:
        linha = (f'{row[0].value}',f'{row[1].value}',f'{row[2].value}',f'{row[3].value}',f'{item}')
        compilado.append(linha)
      except:
        break
  wb.save("Materiamidia.xlsx")

compilacao(dicionario)

