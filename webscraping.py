from scrappings.cnnscrapping import cnnscrapping
from scrappings.epocascrapping import epocascrapping
from scrappings.folhascrapping import folhascrapping
from scrappings.g1scrapping import g1scrapping
from scrappings.vejascrapping import vejascrapping
from openpyxl import load_workbook
from selenium import webdriver
from utils import organizacao_midias 
from utils import compilacao

wb = load_workbook("Materiamidia.xlsx")

options = webdriver.ChromeOptions()
options.add_argument("--headless")
browser = webdriver.Chrome(chrome_options=options)

dicionario = [
        "Banco do Brasil", "Banco do Nordeste", "Banco da Amazônia", "Caixa Econômica Federal", "Banco Central", "Casa da Moeda", "Comissão de Valores Mobiliários", "Susep", "Superintendência de Seguros Privados", "BNDES", "Associação Brasileira de Fundos Garantidores", "BB", "BNB", "BASA", "CEF", "CMB", "ABFG", "AGROS","CAPEF","CAPESESP","CENTRUS","CERES","CIBRIUS","CIFRAO","ELETROS","ELOS","FACHESF","FAPES","FIOPREV","FIPECQ","FUNCEF","FUNPRESP-EXE","FUNPRESP-JUD","FUSESC","GEIPREV","INFRAPREV","NUCLEOS","PETROS","POSTALIS","PREVBEP","PREVDATA","PREVI","PREVINORTE","REAL GRANDEZA","REFER","SAO FRANCISCO","SERPROS","SIAS","CAPAF","PORTUS","TCU"
  ]


for item in dicionario:
  worksheet = organizacao_midias(item)
  try:
    ws = wb[worksheet]
  except:
    wb.create_sheet(worksheet)
    ws = wb[worksheet]
  if item == worksheet or item == "Superintendência de Seguros Privados" or item == "AGROS":
    linha = ("Titulo","Link","Midia","Data")
  ws.append(linha)
  folhascrapping(item,ws,browser)
  cnnscrapping(item,ws,browser)
  epocascrapping(item,ws,browser)
  g1scrapping(item,ws,browser)
  vejascrapping(item,ws,browser)
  wb.save("Materiamidia.xlsx")
compilacao(dicionario,wb)
print("Finish")



